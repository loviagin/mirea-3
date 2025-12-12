#!/usr/bin/env python3
"""
Конвертер CSV в XLSX для Pascal-Legacy модуля.
Обрабатывает типы данных: timestamp, boolean (ИСТИНА/ЛОЖЬ), числа, строки.
"""

import csv
import sys
import os
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_boolean(value: str) -> bool:
    """Парсинг boolean из ИСТИНА/ЛОЖЬ"""
    return value.strip().upper() == 'ИСТИНА'


def parse_timestamp(value: str) -> datetime:
    """Парсинг timestamp в различных форматах"""
    value = value.strip().strip('"')
    # ISO 8601 формат
    for fmt in [
        '%Y-%m-%dT%H:%M:%SZ',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M:%S.%f',
    ]:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse timestamp: {value}")


def convert_csv_to_xlsx(csv_path: str, xlsx_path: str = None):
    """Конвертирует CSV в XLSX с правильной обработкой типов"""
    csv_path = Path(csv_path)
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    
    if xlsx_path is None:
        xlsx_path = csv_path.with_suffix('.xlsx')
    else:
        xlsx_path = Path(xlsx_path)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetry Data"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader)
        
        # Запись заголовков
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Запись данных с обработкой типов
        row_num = 2
        for row in reader:
            for col_idx, (header, value) in enumerate(zip(headers, row), start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                
                # Определение типа по заголовку и значению
                header_lower = header.lower()
                value_stripped = value.strip().strip('"')
                
                if 'timestamp' in header_lower or 'recorded_at' in header_lower or '_at' in header_lower:
                    # TIMESTAMP
                    try:
                        dt = parse_timestamp(value_stripped)
                        cell.value = dt
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                    except (ValueError, AttributeError):
                        cell.value = value_stripped
                
                elif 'is_active' in header_lower or 'boolean' in header_lower or value_stripped in ['ИСТИНА', 'ЛОЖЬ']:
                    # BOOLEAN
                    if value_stripped in ['ИСТИНА', 'ЛОЖЬ']:
                        cell.value = parse_boolean(value_stripped)
                        cell.number_format = 'General'
                    else:
                        cell.value = value_stripped
                
                elif header_lower in ['voltage', 'temp', 'temperature'] or value_stripped.replace('.', '').replace('-', '').isdigit():
                    # Число
                    try:
                        if '.' in value_stripped:
                            cell.value = float(value_stripped)
                        else:
                            cell.value = int(value_stripped)
                    except ValueError:
                        cell.value = value_stripped
                
                else:
                    # Строка
                    cell.value = value_stripped
                
                cell.alignment = Alignment(vertical='center')
            
            row_num += 1
    
    # Автоматическая ширина столбцов
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            if row[0].value:
                max_length = max(max_length, len(str(row[0].value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    wb.save(xlsx_path)
    print(f"✓ Converted {csv_path} -> {xlsx_path}")
    return xlsx_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: csv_to_xlsx.py <csv_file> [xlsx_file]")
        sys.exit(1)
    
    csv_file = sys.argv[1]
    xlsx_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        convert_csv_to_xlsx(csv_file, xlsx_file)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)

